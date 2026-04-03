"""
Извлечение 4 форм МСФО из PDF-отчёта российского банка с помощью MinerU.

Формы:
  1. Отчёт о финансовом положении (баланс)
  2. Отчёт о прибылях и убытках и прочем совокупном доходе
  3. Отчёт о движении денежных средств
  4. Отчёт об изменениях в собственном капитале

Использование:
  python extract_ifrs.py sber.pdf
  python extract_ifrs.py sber.pdf --output output_dir
"""

import sys
import os
import re
import json
import csv
import argparse
import shutil
from pathlib import Path
from html.parser import HTMLParser

# Windows: переключаем stdout на UTF-8, чтобы корректно выводить кириллицу
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ─────────────────────────── MinerU imports ────────────────────────────

from mineru.cli.common import read_fn, do_parse
from mineru.utils.enum_class import MakeMode

# ──────────────────────── Ключевые фразы форм ───────────────────────────

FORM_KEYWORDS: dict[str, list[str]] = {
    "balance_sheet": [
        r"финансово.{0,5}\s*положени",  # финансовом положении / финансового положения
        r"отчет.{0,20}финансово",
        "statement of financial position",
        "balance sheet",
    ],
    "income_statement": [
        r"прибыл.{0,5}\s*и\s*убытк",   # прибылях и убытках / прибыли и убытках
        r"совокупн.{0,5}\s*доход",
        "прочем совокупном",
        "profit or loss",
        "income statement",
        "comprehensive income",
    ],
    "cash_flow": [
        r"движени.{0,10}денежн",        # движении денежных средств
        "cash flow",
        "денежные потоки",
    ],
    "equity_changes": [
        r"изменени.{0,20}собственн",    # изменениях в составе собственных
        r"изменени.{0,10}капитал",
        "changes in equity",
        "собственных средств",
    ],
}

FORM_NAMES_RU: dict[str, str] = {
    "balance_sheet":    "Отчёт о финансовом положении",
    "income_statement": "Отчёт о прибылях и убытках",
    "cash_flow":        "Отчёт о движении денежных средств",
    "equity_changes":   "Отчёт об изменениях в собственном капитале",
}

# ──────────────────────────── HTML → таблица ────────────────────────────

class _TableParser(HTMLParser):
    """
    Парсер HTML-таблиц → список строк (list[list[str]]).
    Обрабатывает colspan/rowspan, <br> внутри ячеек.
    """

    def __init__(self):
        super().__init__()
        self.rows: list[list[str]] = []
        self._cur_row: list[str] = []
        self._cur_cell: list[str] = []
        self._in_cell = False
        self._cur_colspan = 1

    @staticmethod
    def _attr(attrs: list, name: str, default: int = 1) -> int:
        for k, v in attrs:
            if k == name:
                try:
                    return max(1, int(v))
                except (TypeError, ValueError):
                    pass
        return default

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._cur_row = []
        elif tag in ("td", "th"):
            self._cur_cell = []
            self._in_cell = True
            # colspan → add extra empty cells after this one later
            self._cur_colspan = self._attr(attrs, "colspan")
        elif tag == "br" and self._in_cell:
            self._cur_cell.append(" ")

    def handle_endtag(self, tag):
        if tag in ("td", "th"):
            text = "".join(self._cur_cell).strip()
            self._cur_row.append(text)
            # Fill extra columns for colspan
            for _ in range(self._cur_colspan - 1):
                self._cur_row.append("")
            self._in_cell = False
            self._cur_colspan = 1
        elif tag == "tr":
            if self._cur_row:
                self.rows.append(self._cur_row)

    def handle_data(self, data):
        if self._in_cell:
            self._cur_cell.append(data)


def html_to_rows(html: str) -> list[list[str]]:
    parser = _TableParser()
    parser.feed(html)
    return parser.rows


# ───────────────────────── Нормализация числа ───────────────────────────

# Типичные артефакты OCR в числах:
#   «О» вместо «0», «l» вместо «1», слитые тысячные разделители
_DIGIT_FIXES = [
    (re.compile(r'(?<=\d)O(?=\d)'), '0'),   # О → 0 между цифрами
    (re.compile(r'(?<=\d)l(?=\d)'), '1'),   # l → 1 между цифрами
    (re.compile(r'(?<=\d)I(?=\d)'), '1'),   # I → 1 между цифрами
]

# Пробел / тонкая пробельная черта как тысячный разделитель
_THOUSANDS_SEP = re.compile(r'(\d)\s{1,2}(\d{3})(?!\d)')
# Ситуация, когда знак минус — тире/дефис
_MINUS_NORM = re.compile(r'^[–—−‒-]\s*')


def normalize_number(text: str) -> str:
    """
    Нормализует строку с числом:
    - убирает тысячные пробелы (но сохраняет значение)
    - нормализует минус
    - исправляет OCR-артефакты
    Если строка не является числом — возвращает как есть.
    """
    t = text.strip()
    if not t:
        return t

    # Исправление OCR-артефактов
    for pat, repl in _DIGIT_FIXES:
        t = pat.sub(repl, t)

    # Нормализация минуса
    t = _MINUS_NORM.sub('-', t)

    # Убираем тысячные разделители (пробел между группами цифр)
    # Применяем многократно, пока есть совпадения
    prev = None
    while prev != t:
        prev = t
        t = _THOUSANDS_SEP.sub(r'\1\2', t)

    # Проверяем, что это действительно число
    clean = t.replace(',', '.').replace(' ', '')
    try:
        float(clean)
        return t          # числовая строка — возвращаем нормализованную
    except ValueError:
        return text.strip()  # не число — оригинал


def clean_cell(text: str) -> str:
    """Очистка ячейки: убирает лишние переносы, нормализует пробелы."""
    # Убираем мягкий перенос и другие невидимые символы
    text = text.replace('\u00ad', '').replace('\u200b', '').replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return normalize_number(text)


# ──────────────────────── Определение формы ────────────────────────────

def _text_matches(text: str, form_key: str) -> bool:
    t = text.lower()
    for kw in FORM_KEYWORDS[form_key]:
        if re.search(kw, t, re.IGNORECASE):
            return True
    return False


def detect_form_from_caption(caption: str) -> str | None:
    """
    Определяем форму только по caption таблицы (более надёжно).
    Приоритет: caption проверяется первым.
    """
    if not caption:
        return None
    for key in ("cash_flow", "equity_changes", "income_statement", "balance_sheet"):
        if _text_matches(caption, key):
            return key
    return None


def detect_form(context_texts: list[str]) -> str | None:
    """
    По списку текстовых блоков рядом с таблицей определяем,
    к какой форме она относится.
    Возвращает ключ формы или None.
    """
    combined = " ".join(context_texts)
    # Приоритет: сначала более специфичные
    for key in ("cash_flow", "equity_changes", "income_statement", "balance_sheet"):
        if _text_matches(combined, key):
            return key
    return None


# ───────────────────── Парсинг content_list.json ───────────────────────

_NOTES_RE = re.compile(
    r'^\s*(примечани|поясне|note\s*\d)',
    re.IGNORECASE,
)

# Страницы сверх этого предела обрабатываются только если есть точный caption
_MAX_MAIN_PAGES = 30  # типично формы МСФО занимают первые 10-20 страниц

# Слова, с которых начинаются самостоятельные строки (итоги, подстатьи),
# даже если первая буква строчная — их нельзя склеивать с предыдущей строкой
_STANDALONE_RE = re.compile(
    r'^(итого|всего|в\s+том\s+числе|в\s+т\.?\s*ч\.?|из\s+них|из\s+которых'
    r'|включая|прочие|иные|справочно|в\s+том\s+числе|нераспределённ|нераспределен)',
    re.IGNORECASE,
)

# Маркеры строк оглавления: ". 6", ". .12", ".. 33" — номера страниц
_TOC_RE = re.compile(r'\.\s*\.?\s*\d{1,3}(?:\s|$)', re.MULTILINE)

# Жёсткое начало раздела примечаний
# Отличает реальный раздел примечаний от ссылок в оглавлении/аудиторском тексте
_NOTES_HARD_RE = re.compile(
    r'(?:'
    # «ПРИМЕЧАНИЯ К КОНСОЛИДИРОВАННОЙ/ФИНАНСОВОЙ/СВОДНОЙ...»
    r'ПРИМЕЧАНИ[А-Я]\s+К\s+(?:КОНСОЛИДИРОВАННОЙ|ФИНАНСОВОЙ|СВОДНОЙ|ОТДЕЛЬНОЙ)'
    r'|'
    # «Примечания к финансовой отчётности за ... 2024»
    r'примечани.{2,60}(?:20\d{2}|декабр|финансов|отчётност|отчетност)'
    r'|'
    # «Пояснения к финансовой отчётности»
    r'поясне.{2,60}(?:финансов|отчётност|отчетност)'
    r'|'
    # English: Notes to the (consolidated/financial) statements
    r'notes\s+to\s+the\s+(?:consolidated|financial|separate)\s+(?:financial\s+)?statements'
    r')',
    re.IGNORECASE,
)

# Паттерны единиц измерения для определения масштаба чисел
_UNIT_PATTERNS: list[tuple[re.Pattern, int, str]] = [
    (re.compile(r'(?:в\s+)?(?:тысяч.{0,5}\s*рубл|тыс\.?\s*руб\.?|в\s+тыс\.)', re.IGNORECASE),
     1_000, "тыс. руб."),
    (re.compile(r'(?:в\s+)?(?:миллион.{0,5}\s*рубл|млн\.?\s*руб\.?)', re.IGNORECASE),
     1_000_000, "млн. руб."),
    (re.compile(r'(?:в\s+)?(?:миллиард.{0,5}\s*рубл|млрд\.?\s*руб\.?)', re.IGNORECASE),
     1_000_000_000, "млрд. руб."),
]

# Максимальная длина текстового блока, из которого можно определить форму.
# Длинные абзацы (аудиторское заключение, примечания) пропускаются.
_MAX_HEADING_LEN = 180


def extract_tables_from_content_list(
    content_list_path: str,
) -> tuple[dict[str, list[list[list[str]]]], dict[str, str]]:
    """
    Читает content_list.json и собирает таблицы по 4 формам.

    Стратегия:
    1. Caption таблицы имеет высший приоритет.
    2. Контекст (последний заголовок формы в тексте) — fallback для таблиц
       без caption. Контекст устанавливается ТОЛЬКО из коротких текстовых
       блоков (≤ 180 символов), которые не являются строками оглавления.
    3. Как только начинается настоящий раздел примечаний, дальнейший
       контекстный поиск прекращается — принимаются только таблицы с caption.
    4. Блоки после страницы _MAX_MAIN_PAGES тоже принимаются только
       при наличии точного caption (дополнительная защита).
    5. Попутно определяется единица измерения для каждой формы.

    Возвращает: (form_tables_dict, units_dict)
    """
    with open(content_list_path, encoding="utf-8") as f:
        blocks = json.load(f)

    result: dict[str, list[list[list[str]]]] = {k: [] for k in FORM_KEYWORDS}
    # Единицы измерения: form_key → строка ("тыс. руб." / "млн. руб." / "")
    units: dict[str, str] = {k: "" for k in FORM_KEYWORDS}

    current_form: str | None = None
    notes_started: bool = False  # True — находимся в разделе примечаний
    current_page: int  = 0       # текущий номер страницы (0-based)

    for block in blocks:
        btype    = block.get("type", "")
        current_page = block.get("page_idx", current_page)

        # За пределами основных страниц и без точного caption — пропускаем
        beyond_limit = current_page >= _MAX_MAIN_PAGES

        if btype == "text":
            txt = block.get("text", "").strip()
            if not txt:
                continue

            # Жёсткое начало примечаний → полная остановка контекста
            if not notes_started and _NOTES_HARD_RE.search(txt) and len(txt) < 350:
                notes_started = True
                current_form  = None
                continue

            # В разделе примечаний и за лимитом страниц — не обновляем контекст
            if notes_started or beyond_limit:
                continue

            # Строки оглавления — пропускаем
            if _TOC_RE.search(txt):
                continue

            # Мягкий сброс: «Примечание/Примечания» в начале блока
            if _NOTES_RE.search(txt):
                current_form = None
                continue

            # Определяем единицы измерения из текстовых блоков
            for upat, _mult, ulabel in _UNIT_PATTERNS:
                if upat.search(txt):
                    # Присваиваем текущей форме, если она определена
                    if current_form and not units[current_form]:
                        units[current_form] = ulabel
                    break

            # Форму определяем только из коротких заголовков
            if len(txt) <= _MAX_HEADING_LEN:
                for key in ("cash_flow", "equity_changes", "income_statement", "balance_sheet"):
                    if _text_matches(txt, key):
                        current_form = key
                        # Попробуем определить единицу из этого же блока
                        for upat, _mult, ulabel in _UNIT_PATTERNS:
                            if upat.search(txt):
                                if not units[current_form]:
                                    units[current_form] = ulabel
                                break
                        break

        elif btype == "table":
            html = block.get("table_body") or block.get("html") or ""
            if not html:
                continue

            rows = html_to_rows(html)
            if not rows:
                continue

            caption = block.get("table_caption", [])
            caption_text = " ".join(caption) if isinstance(caption, list) else str(caption)

            # Caption — всегда (даже в зоне примечаний или за лимитом страниц)
            form_key = detect_form_from_caption(caption_text)

            # Контекст — только вне раздела примечаний и в пределах лимита
            if not form_key and not notes_started and not beyond_limit:
                form_key = current_form

            if form_key:
                result[form_key].append(rows)
                # Если единица ещё не определена, ищем в заголовке таблицы
                if not units[form_key] and caption_text:
                    for upat, _mult, ulabel in _UNIT_PATTERNS:
                        if upat.search(caption_text):
                            units[form_key] = ulabel
                            break

    return result, units


# ─────────────────────── Склейка таблиц (pagination) ──────────────────

def merge_table_pages(tables: list[list[list[str]]]) -> list[list[str]]:
    """
    Склеивает список таблиц одной формы (разбитых по страницам).
    Дублирующийся заголовок (первая строка) при склейке пропускается.
    """
    if not tables:
        return []
    merged: list[list[str]] = [list(row) for row in tables[0]]
    if len(tables) == 1:
        return merged

    header = tables[0][0] if tables[0] else []

    for tbl in tables[1:]:
        if not tbl:
            continue
        # Если первая строка совпадает с заголовком — пропускаем
        start = 1 if (tbl[0] == header or _rows_similar(tbl[0], header)) else 0
        merged.extend(list(row) for row in tbl[start:])

    return merged


def _rows_similar(a: list[str], b: list[str]) -> bool:
    """Проверяет, похожи ли две строки (заголовки) по большинству ячеек."""
    if not a or not b:
        return False
    matches = sum(1 for x, y in zip(a, b) if x.strip().lower() == y.strip().lower())
    return matches / max(len(a), len(b)) > 0.6


# ──────────────────────── Исправление разрывов строк ─────────────────────

def fix_line_breaks_in_table(rows: list[list[str]]) -> list[list[str]]:
    """
    Исправляет разрывы строк внутри ячеек таблиц.

    Строка считается продолжением предыдущей, если:
    - её первая непустая ячейка начинается со строчной буквы, И
    - в колонках с данными (1+) нет конфликта: у обеих строк одновременно
      заполнена одна и та же колонка (это означало бы две отдельные строки).

    При склейке:
    - текст продолжения дописывается к правильной колонке предыдущей строки
      (той, где найдена первая непустая ячейка продолжения)
    - данные из остальных колонок переносятся ТОЛЬКО в пустые колонки
      предыдущей строки (не перезаписывают существующие значения)

    Пример:
      Было:
        ["Процентные доходы, рассчитанные по методу эффективной", "", "", ""]
        ["ставки", "463,3", "269,0", "-20,9%", "24,0%"]
        ["Чистые комиссионные доходы", "", "", "", ""]
      Стало:
        ["Процентные доходы, рассчитанные по методу эффективной ставки", "463,3", "269,0", "-20,9%", "24,0%"]
        ["Чистые комиссионные доходы", "", "", "", ""]
    """
    if not rows:
        return []

    result: list[list[str]] = []

    for row in rows:
        if not row:
            continue

        # Находим первую непустую ячейку и её индекс
        first_idx = -1
        first_val = ""
        for i, cell in enumerate(row):
            stripped = cell.strip()
            if stripped:
                first_idx = i
                first_val = stripped
                break

        # Определяем, является ли строка продолжением предыдущей.
        # Условия для склейки (все должны выполняться):
        #   1. Есть предыдущая строка
        #   2. Первый непустой символ — строчная буква
        #   3. Текст не начинается с известного «самостоятельного» слова
        #      (итого / в том числе / из них / прочие / ...)
        #   4. Предыдущая строка имеет текст в колонке 0 (это строка-метка,
        #      а не строка-заголовок с годами или пустая строка)
        #   5. Нет конфликта данных: в одних и тех же колонках (1+)
        #      не могут быть одновременно непустые значения у обеих строк
        is_continuation = False
        if (
            result
            and first_val
            and first_val[0].islower()
            and not _STANDALONE_RE.match(first_val)
        ):
            prev_row = result[-1]
            prev_label = prev_row[0].strip() if prev_row else ""

            # Предыдущая строка должна иметь текстовую метку в колонке 0
            if prev_label:
                conflict = False
                for i in range(1, max(len(row), len(prev_row))):
                    cur_val  = row[i].strip()      if i < len(row)      else ""
                    prev_val = prev_row[i].strip() if i < len(prev_row) else ""
                    if cur_val and prev_val:
                        conflict = True
                        break
                if not conflict:
                    is_continuation = True

        if is_continuation:
            prev_row = result[-1]

            # Расширяем предыдущую строку до размера текущей, если нужно
            while len(prev_row) < len(row):
                prev_row.append("")

            # Дописываем текст продолжения в правильную колонку
            if first_idx >= 0:
                if first_idx < len(prev_row) and prev_row[first_idx].strip():
                    prev_row[first_idx] = prev_row[first_idx].rstrip() + " " + first_val
                elif first_idx < len(prev_row):
                    prev_row[first_idx] = first_val

            # Переносим данные из остальных колонок ТОЛЬКО в пустые ячейки
            for i in range(len(row)):
                if i == first_idx:
                    continue
                val = row[i].strip()
                if not val:
                    continue
                if i < len(prev_row):
                    if not prev_row[i].strip():
                        prev_row[i] = val
                else:
                    prev_row.append(val)
        else:
            result.append(list(row))

    return result


# ──────────────────── Исправление склеенных ячеек ───────────────────────

def fix_split_cells(rows: list[list[str]]) -> list[list[str]]:
    """
    Исправляет артефакт MinerU, когда два числа из соседних строк HTML
    попадают в одну ячейку ("3 901,6 1707,5").

    Алгоритм: если ячейка содержит два валидных числа (N1 N2) и ячейка
    предыдущей строки в той же колонке пуста — перемещаем N1 туда.

    Обрабатывается только колонки 1+ (не колонка с подписью/меткой).
    """
    for ri in range(1, len(rows)):
        row  = rows[ri]
        prev = rows[ri - 1]
        for ci in range(1, len(row)):          # только числовые колонки
            val = row[ci].strip()
            if not val:
                continue
            parts = val.split()
            if len(parts) < 2:
                continue
            # Перебираем возможные точки разбиения
            for split_pt in range(1, len(parts)):
                left  = " ".join(parts[:split_pt])
                right = " ".join(parts[split_pt:])
                if _to_number(left) is not None and _to_number(right) is not None:
                    # Предыдущая ячейка в той же колонке должна быть пустой
                    prev_cell = prev[ci].strip() if ci < len(prev) else ""
                    if not prev_cell:
                        # Расширяем предыдущую строку при необходимости
                        while len(prev) <= ci:
                            prev.append("")
                        prev[ci] = left
                        row[ci]  = right
                    break   # нашли разбиение — следующие не проверяем
    return rows


# ──────────────────────────── Запись CSV ────────────────────────────────

def write_csv(rows: list[list[str]], path: str) -> None:
    """Записывает таблицу в CSV (UTF-8 с BOM для совместимости с Excel)."""
    max_cols = max((len(r) for r in rows), default=0)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            # Выравниваем строки по максимальной ширине
            padded = row + [""] * (max_cols - len(row))
            cleaned = [clean_cell(c) for c in padded]
            writer.writerow(cleaned)


# ──────────────────────────── Запись Excel ──────────────────────────────

def _to_number(val: str):
    """
    Пробует преобразовать строку в число. Возвращает int/float или None.
    Обрабатывает:
      - Отрицательные в скобках: (1 234,5) → -1234.5
      - Двойные скобки-артефакты OCR: (3 555,1)) → -3555.1
      - Тысячные разделители (пробел, тонкая пробельная черта)
      - Запятую как десятичный разделитель
      - Знак % (возвращает None — процент остаётся текстом)
      - Строки с несколькими числами (возвращает None — нельзя выбрать одно)
    """
    v = val.strip()
    if not v or v in ("-", "—", "–", "−", "н/д", "н.д.", "x", "х",
                      "н.п.", "н/п", "H.П.", "Н.П.", "—", "н.п."):
        return None
    # Процент — оставляем текстом
    if "%" in v:
        return None
    # Если строка содержит несколько значений (артефакт слияния строк) → None
    # Признак: два числа через пробел: "3 979,4 122,5" после нормализации → два числа
    # Определяем ДО удаления пробелов-разделителей тысяч
    # Грубая проверка: после убирания тысячных групп остаётся пробел внутри числа
    v_check = re.sub(r'(\d)\s(\d{3})(?!\d)', r'\1\2', v)  # убираем тысячные разделители
    v_check = v_check.replace(",", ".").strip("()")
    # Если в строке несколько разделённых пробелом числоподобных частей → пропускаем
    parts = v_check.split()
    if len(parts) > 1:
        numeric_parts = sum(1 for p in parts if re.fullmatch(r'-?[\d.]+', p.strip("(),")))
        if numeric_parts > 1:
            return None  # несколько чисел — нельзя выбрать одно

    # Отрицательное число в скобках (с возможными артефактами `))`)
    # Удаляем все открывающие/закрывающие скобки с краёв и определяем знак
    negative = "(" in v
    v = v.replace("(", "").replace(")", "")

    # Убираем пробелы-разделители тысяч и тонкие пробелы
    v = v.replace(" ", "").replace("\u202f", "").replace("\xa0", "")
    # Запятая → точка
    v = v.replace(",", ".")
    # Оставляем только цифры, точку и минус
    v = re.sub(r"[^\d.\-]", "", v)
    if not v or v in (".", "-"):
        return None
    try:
        num = float(v)
        num = -num if negative else num
        # Возвращаем int, если число целое
        return int(num) if num == int(num) else round(num, 4)
    except ValueError:
        return None


# ──────────────────────── Определение типа строки ────────────────────────

_CAPS_ALPHA_RE = re.compile(r'^[А-ЯЁA-Z\s\-/()]{4,}$')   # строка из заглавных букв
_TOTAL_RE      = re.compile(r'^(итого|всего|total)\b', re.IGNORECASE)
_COLHDR_YEAR   = re.compile(r'\b(20\d{2}|прим\.?|note)\b', re.IGNORECASE)


def _classify_row(row: list[str]) -> str:
    """
    Возвращает тип строки для стилизации Excel:
      'form_header' — заголовок секции формы (АКТИВЫ, ОБЯЗАТЕЛЬСТВА …)
      'col_header'  — заголовок колонок (Прим., 2024г., 2023г. …)
      'total'       — итоговая строка (Итого, Всего …)
      'data'        — обычная строка с данными
    """
    label = row[0].strip() if row else ""
    # Пустой label и колонки содержат годы → заголовок колонок
    if not label and any(_COLHDR_YEAR.search(c) for c in row[1:] if c.strip()):
        return "col_header"
    # Label состоит из заглавных букв (ALL-CAPS секция)
    if label and _CAPS_ALPHA_RE.match(label) and not any(c.strip() for c in row[1:]):
        return "form_header"
    # Строка-итог
    if _TOTAL_RE.match(label):
        return "total"
    # Заголовок колонок: label == "Прим." / пустой + в колонках годы
    if label in ("Прим.", "Прим") or _COLHDR_YEAR.match(label):
        data_cells = [c for c in row[1:] if c.strip()]
        if not data_cells or all(_COLHDR_YEAR.search(c) for c in data_cells):
            return "col_header"
    return "data"


def write_excel(
    processed_forms: dict[str, list[list[str]]],
    path: str,
) -> None:
    """
    Записывает все найденные формы на один лист Excel.
    Каждая форма предваряется тёмно-синим заголовком.
    Стили строк:
      - Секционные заголовки (АКТИВЫ, ОБЯЗАТЕЛЬСТВА) — серый фон, жирный
      - Заголовки колонок (Прим., 2024г.) — голубой фон, жирный
      - Итоговые строки (Итого, Всего) — светло-серый фон, жирный
      - Числовые ячейки — правое выравнивание, числовой тип
    Требует: pip install openpyxl
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [!] openpyxl не установлен, Excel не создан. Выполните: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "МСФО"

    # ── Палитра стилей ───────────────────────────────────────────────────
    S = {
        "form_title": {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill("solid", fgColor="1F4E79"),
            "height": 22,
        },
        "col_header": {
            "font": Font(bold=True, size=10),
            "fill": PatternFill("solid", fgColor="BDD7EE"),
            "height": 16,
        },
        "form_header": {
            "font": Font(bold=True, size=10),
            "fill": PatternFill("solid", fgColor="D9D9D9"),
            "height": 15,
        },
        "total": {
            "font": Font(bold=True, size=10),
            "fill": PatternFill("solid", fgColor="F2F2F2"),
            "height": 15,
        },
        "data": {
            "font": Font(size=10),
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "height": 15,
        },
    }

    # Ширина столбцов: col 1 (labels) = 55, col 2+ (данные) = 14
    LABEL_COL_WIDTH = 55
    DATA_COL_WIDTH  = 14

    cur_row = 1
    max_data_col = 1  # Максимальная колонка среди всех форм

    for form_key, rows in processed_forms.items():
        if not rows:
            continue

        form_name = FORM_NAMES_RU[form_key]
        max_cols  = max((len(r) for r in rows), default=1)
        max_data_col = max(max_data_col, max_cols)

        # ── Заголовок формы ─────────────────────────────────────────────
        ws.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row,   end_column=max(max_cols, 2),
        )
        hcell = ws.cell(row=cur_row, column=1, value=form_name)
        hcell.font      = S["form_title"]["font"]
        hcell.fill      = S["form_title"]["fill"]
        hcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur_row].height = S["form_title"]["height"]
        cur_row += 1

        # ── Строки таблицы ───────────────────────────────────────────────
        for row in rows:
            padded   = row + [""] * (max_cols - len(row))
            cleaned  = [clean_cell(c) for c in padded]
            row_type = _classify_row(cleaned)
            style    = S[row_type]

            for col_idx, val in enumerate(cleaned, 1):
                # Числовое значение только для колонок с данными (2+)
                num = _to_number(val) if col_idx > 1 else None

                cell = ws.cell(
                    row=cur_row,
                    column=col_idx,
                    value=num if num is not None else (val or None),
                )
                cell.font = style["font"]
                cell.fill = style["fill"]

                if row_type in ("col_header", "form_header"):
                    cell.alignment = Alignment(
                        horizontal="center" if col_idx > 1 else "left",
                        vertical="center",
                        wrap_text=False,
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="right" if num is not None else "left",
                        vertical="center",
                        indent=1 if col_idx == 1 else 0,
                    )

            ws.row_dimensions[cur_row].height = style["height"]
            cur_row += 1

        cur_row += 2  # пустые строки между формами

    # ── Ширина колонок ───────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = LABEL_COL_WIDTH
    for col in range(2, max_data_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = DATA_COL_WIDTH

    # ── Закрепляем первую колонку ────────────────────────────────────────
    ws.freeze_panes = "B1"

    wb.save(path)


# ────────────────────────────── Main ────────────────────────────────────

def run(pdf_path: str, output_dir: str) -> None:
    pdf_path = str(Path(pdf_path).resolve())
    pdf_name = Path(pdf_path).stem
    output_dir = str(Path(output_dir).resolve())

    mineru_out = os.path.join(output_dir, "_mineru_raw")
    os.makedirs(mineru_out, exist_ok=True)

    # Кэш: если content_list.json уже создан — пропускаем долгий OCR
    expected_cl = os.path.join(mineru_out, pdf_name, "txt", f"{pdf_name}_content_list.json")
    cached_matches = list(Path(mineru_out).rglob("*_content_list.json"))
    content_list_path = str(cached_matches[0]) if cached_matches else None

    if content_list_path and os.path.exists(content_list_path):
        print(f"[1/3] Кэш найден, MinerU пропускается: {content_list_path}")
    else:
        print(f"[1/3] Запуск MinerU на '{pdf_path}' (может занять несколько минут) ...")
        pdf_bytes = read_fn(pdf_path)

        do_parse(
            output_dir=mineru_out,
            pdf_file_names=[pdf_name],
            pdf_bytes_list=[pdf_bytes],
            p_lang_list=["east_slavic"],   # русский / кириллица
            backend="pipeline",
            parse_method="txt",            # текстовый PDF (не OCR)
            formula_enable=False,          # формулы нам не нужны
            table_enable=True,
            f_draw_layout_bbox=False,
            f_draw_span_bbox=False,
            f_dump_md=False,
            f_dump_middle_json=False,
            f_dump_model_output=False,
            f_dump_orig_pdf=False,
            f_dump_content_list=True,      # нужен content_list.json
            f_make_md_mode=MakeMode.MM_MD,
        )

        # Находим content_list.json
        content_list_path = expected_cl
        if not os.path.exists(content_list_path):
            found = list(Path(mineru_out).rglob("*_content_list.json"))
            if not found:
                raise FileNotFoundError(
                    f"content_list.json не найден в {mineru_out}. "
                    "Возможно, MinerU не смог обработать PDF."
                )
            content_list_path = str(found[0])

    print(f"[2/3] Разбор content_list.json: {content_list_path}")
    form_tables, form_units = extract_tables_from_content_list(content_list_path)

    print(f"[3/3] Запись результатов в '{output_dir}' ...")
    os.makedirs(output_dir, exist_ok=True)

    found_forms: list[str] = []
    missing_forms: list[str] = []
    processed: dict[str, list[list[str]]] = {}  # form_key → готовые строки

    for form_key, tables in form_tables.items():
        name_ru = FORM_NAMES_RU[form_key]
        if not tables:
            print(f"  [!] НЕ НАЙДЕНО: {name_ru}")
            missing_forms.append(form_key)
            processed[form_key] = []
            continue

        merged = merge_table_pages(tables)
        merged = fix_split_cells(merged)
        merged = fix_line_breaks_in_table(merged)
        processed[form_key] = merged

        csv_path = os.path.join(output_dir, f"{form_key}.csv")
        write_csv(merged, csv_path)
        print(f"  [+] {name_ru} -> {csv_path}  ({len(merged)} строк, {len(tables)} частей)")
        found_forms.append(form_key)

    # Сохраняем метаданные единиц измерения
    units_path = os.path.join(output_dir, "form_units.json")
    with open(units_path, "w", encoding="utf-8") as fu:
        json.dump(form_units, fu, ensure_ascii=False, indent=2)
    if any(v for v in form_units.values()):
        detected = {k: v for k, v in form_units.items() if v}
        print(f"  [+] Единицы измерения: {detected}")
    else:
        print("  [!] Единицы измерения не определены (возможно, указаны в заголовке страницы вне таблиц)")

    # Excel: все формы на одном листе
    xlsx_path = os.path.join(output_dir, f"{pdf_name}_ifrs.xlsx")
    write_excel(processed, xlsx_path)
    if os.path.exists(xlsx_path):
        print(f"  [+] Excel -> {xlsx_path}")

    print()
    if missing_forms:
        print(
            "Не найдены формы:", ", ".join(FORM_NAMES_RU[k] for k in missing_forms)
        )
        print("Совет: если PDF на основе OCR, повторите с parse_method='ocr'.")
    if found_forms:
        print(f"Готово! Найдено {len(found_forms)}/4 форм. Результаты в: {output_dir}")

    # Удаляем временные файлы MinerU (images и т.п.) — оставляем только CSV
    # Раскомментируйте, если не нужны промежуточные файлы:
    # shutil.rmtree(mineru_out, ignore_errors=True)


def main():
    parser = argparse.ArgumentParser(description="Извлечение 4 форм МСФО из PDF банковского отчёта")
    parser.add_argument("pdf", help="Путь к PDF-файлу (например, sber.pdf)")
    parser.add_argument(
        "--output", "-o",
        default="ifrs_output",
        help="Папка для CSV-файлов (по умолчанию: ifrs_output)",
    )
    args = parser.parse_args()

    run(args.pdf, args.output)


if __name__ == "__main__":
    main()