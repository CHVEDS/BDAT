"""
МСФО-экстрактор — Streamlit-интерфейс
Корпоративный стиль (синяя гамма)

Запуск:
    streamlit run app_ifrs.py
"""

import io
import os
import sys
import contextlib
import tempfile
import zipfile
from pathlib import Path
from collections import defaultdict

import streamlit as st
import pandas as pd

import extract_ifrs
import metadata_extractor as meta_ext

# ─────────────────────────── Стили ──────────────────────────────────────

C_BLUE      = "#002B6E"   # основной синий (тёмный) — белый текст поверх
C_BLUE2     = "#0047AB"   # средний синий
C_BLUE_LT   = "#C8DAEA"   # светло-голубой фон (только для акцентов)
C_BORDER    = "#9DBDD6"   # граница элементов
C_BG        = "#B0E0E6"   # фон страницы
C_WHITE     = "#4682B4"   # фон карточек (стальной синий)
C_TEXT      = "#FFFFFF"   # основной текст (белый)
C_TEXT_2    = "#2D3748"   # вторичный текст — тёмно-серый (8:1 на белом)
C_TEXT_MUTE = "#4A5568"   # приглушённый текст (7:1 на белом)
C_GREEN     = "#145A32"   # зелёный текст (11:1 на зелёном фоне)
C_GREEN_BG  = "#C3E6CB"   # зелёный фон (для зелёного текста)
C_RED       = "#7B1E1E"   # красный текст (10:1 на красном фоне)
C_RED_BG    = "#F5C6CB"   # красный фон
C_YELLOW_BG = "#FFEEBA"   # жёлтый фон
C_YELLOW    = "#6B4A00"   # тёмно-жёлтый текст (9:1 на жёлтом фоне)

CSS = f"""
<style>
/* ════════════════════════════════════════════════════════════════════
   ВАЖНО: НЕ переопределяем font-family на нативных элементах Streamlit
   (span, p, label, div) — это ломает Material Icons (иконки становятся
   текстом вида "keyboard_double_arrow_left").
   font-family назначаем ТОЛЬКО нашим кастомным HTML-классам.
   ════════════════════════════════════════════════════════════════════ */

/* ── Фон приложения ── */
.stApp {{
    background-color: {C_BG};
}}

/* ── Заголовки (h1-h4) ── только цвет, без font-family ── */
.stApp h1, .stApp h2, .stApp h3, .stApp h4 {{
    color: {C_BLUE} !important;
}}

/* ── Текст в markdown-блоках ── */
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li,
[data-testid="stMarkdownContainer"] strong,
[data-testid="stText"] {{
    color: {C_TEXT} !important;
}}

/* ── Подписи виджетов (selectbox, radio, multiselect, file) ── */
[data-testid="stWidgetLabel"] > div > p {{
    color: {C_TEXT} !important;
    font-weight: 600;
}}

/* ── Текст внутри radio / checkbox ── */
.stRadio div[role="radiogroup"] label > div > p,
.stCheckbox label > div > p {{
    color: {C_TEXT} !important;
}}

/* ── Sidebar: только фон, без перегрузки span/div (ломает иконки) ── */
section[data-testid="stSidebar"] {{
    background: {C_WHITE};
    border-right: 2px solid {C_BLUE_LT};
}}
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {{
    color: #FFFFFF !important;
}}
/* Текстовые параграфы сайдбара — без span и label */
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {{
    color: {C_TEXT} !important;
}}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {{
    gap: 4px;
    border-bottom: 2px solid {C_BLUE_LT};
    background: transparent;
}}
.stTabs [data-baseweb="tab"] {{
    font-size: 13px;
    font-weight: 600;
    color: {C_TEXT_2} !important;
    background: {C_BG};
    border: 1px solid {C_BORDER};
    border-bottom: none;
    border-radius: 6px 6px 0 0;
    padding: 6px 14px;
}}
.stTabs [aria-selected="true"] {{
    color: #FFFFFF !important;
    border-color: {C_BLUE} !important;
    background: {C_BLUE} !important;
}}
.stTabs [data-baseweb="tab-highlight"] {{
    background-color: {C_BLUE} !important;
}}
.stTabs [data-baseweb="tab-panel"] {{
    background: {C_WHITE};
    border: 1px solid {C_BLUE_LT};
    border-top: none;
    border-radius: 0 0 8px 8px;
    padding: 14px;
}}

/* ── Кнопки действий ── */
div.stButton > button {{
    border-radius: 6px;
    font-weight: 600;
    font-size: 13px;
    padding: 8px 20px;
    transition: background-color 0.2s;
}}
div.stButton > button[kind="primary"] {{
    background: {C_BLUE} !important;
    color: #FFFFFF !important;
    border: none;
}}
div.stButton > button[kind="primary"]:hover {{
    background: {C_BLUE2} !important;
}}
div.stButton > button[kind="secondary"] {{
    background: {C_WHITE} !important;
    color: #FFFFFF !important;
    border: 2px solid {C_BLUE_LT};
}}

/* ── Кнопки скачивания ── */
div[data-testid="stDownloadButton"] > button {{
    background: {C_BLUE} !important;
    color: #FFFFFF !important;
    border: none;
    border-radius: 6px;
    font-weight: 600;
    font-size: 13px;
    padding: 8px 18px;
}}
div[data-testid="stDownloadButton"] > button:hover {{
    background: {C_BLUE2} !important;
}}

/* ── Таблицы ── */
.stDataFrame {{
    border: 1px solid {C_BORDER};
    border-radius: 6px;
    overflow: hidden;
}}

/* ── Сообщения ── */
div[data-testid="stAlert"] {{
    border-radius: 6px;
}}

/* ══════════════════════════════════════════════════════════════════
   EXPANDER — кнопка разворота
   Streamlit рендерит expander через <details>/<summary>.
   Стрелка — SVG, а не Material Icon → override через svg path безопасен.
   ══════════════════════════════════════════════════════════════════ */
[data-testid="stExpander"] details summary {{
    background: {C_WHITE};
    border: 1px solid {C_BORDER};
    border-radius: 6px;
    padding: 10px 14px;
    cursor: pointer;
}}
[data-testid="stExpander"] details[open] summary {{
    border-bottom-left-radius: 0;
    border-bottom-right-radius: 0;
}}
/* SVG-стрелка экспандера */
[data-testid="stExpander"] details summary svg {{
    stroke: {C_BLUE} !important;
}}
[data-testid="stExpander"] details summary svg path {{
    stroke: {C_BLUE} !important;
}}

/* ══════════════════════════════════════════════════════════════════
   SIDEBAR TOGGLE — кнопка сворачивания/разворачивания панели.
   Streamlit использует Material Icons (текстовые лигатуры) — НЕ трогаем
   font-family и color у вложенных span, иначе иконка отображается
   как "keyboard_double_arrow_left" вместо символа.
   Безопасно менять только background и border у самой кнопки.
   ══════════════════════════════════════════════════════════════════ */
[data-testid="stSidebarCollapseButton"] button,
[data-testid="stSidebarCollapsedControl"] button {{
    background: {C_WHITE} !important;
    border: 1.5px solid {C_BORDER} !important;
    border-radius: 8px !important;
}}

/* ── Разделитель ── */
hr {{ border-top: 2px solid {C_BLUE_LT}; margin: 20px 0; }}

/* ════════════════════════════════════════════════════════════════════
   НАШИ КАСТОМНЫЕ HTML-ЭЛЕМЕНТЫ — здесь font-family безопасен,
   т.к. они не содержат Material Icons.
   ════════════════════════════════════════════════════════════════════ */

/* ── Шапка приложения ── */
.app-header {{
    background: {C_BLUE};
    padding: 20px 32px 16px;
    border-radius: 0 0 10px 10px;
    margin-bottom: 24px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.22);
    font-family: "Arial", "Helvetica Neue", sans-serif;
}}
.app-header h1 {{
    color: #FFFFFF !important;
    font-size: 20px;
    font-weight: 700;
    margin: 0;
    line-height: 1.3;
}}
.app-header p {{
    color: {C_BLUE_LT} !important;
    font-size: 13px;
    margin: 4px 0 0;
}}

/* ── Карточки ── */
.card {{
    background: {C_WHITE};
    border: 1px solid {C_BORDER};
    border-radius: 8px;
    padding: 18px 20px;
    margin-bottom: 12px;
    box-shadow: 0 1px 4px rgba(0,43,110,0.12);
    font-family: "Arial", "Helvetica Neue", sans-serif;
}}
.card-header {{
    font-size: 13px;
    font-weight: 700;
    color: #FFFFFF !important;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 10px;
    padding-bottom: 6px;
    border-bottom: 2px solid {C_BLUE_LT};
}}

/* ── Стат-плитки ── */
.stat-tile {{
    background: {C_WHITE};
    border: 1px solid {C_BORDER};
    border-top: 3px solid {C_BLUE2};
    border-radius: 8px;
    padding: 14px 16px;
    text-align: center;
    box-shadow: 0 1px 3px rgba(0,43,110,0.1);
    font-family: "Arial", "Helvetica Neue", sans-serif;
}}
.stat-tile .val {{
    font-size: 28px;
    font-weight: 700;
    color: #FFFFFF !important;
    line-height: 1.1;
}}
.stat-tile .lbl {{
    font-size: 12px;
    font-weight: 600;
    color: {C_BLUE_LT} !important;
    margin-top: 3px;
}}
.stat-tile .sub {{
    font-size: 11px;
    margin-top: 6px;
}}

/* ── Бейджи ── */
.badge-ok {{
    background: {C_GREEN_BG};
    color: {C_GREEN} !important;
    border-radius: 10px;
    padding: 2px 9px;
    font-size: 11px;
    font-weight: 700;
    white-space: nowrap;
    font-family: "Arial", sans-serif;
}}
.badge-err {{
    background: {C_RED_BG};
    color: {C_RED} !important;
    border-radius: 10px;
    padding: 2px 9px;
    font-size: 11px;
    font-weight: 700;
    white-space: nowrap;
    font-family: "Arial", sans-serif;
}}
.badge-warn {{
    background: {C_YELLOW_BG};
    color: {C_YELLOW} !important;
    border-radius: 10px;
    padding: 2px 9px;
    font-size: 11px;
    font-weight: 700;
    white-space: nowrap;
    font-family: "Arial", sans-serif;
}}

/* ── Полоса заголовка формы ── */
.form-bar {{
    background: {C_BLUE};
    color: #FFFFFF !important;
    font-size: 13px;
    font-weight: 600;
    padding: 8px 14px;
    border-radius: 6px 6px 0 0;
    margin-bottom: 0;
    font-family: "Arial", sans-serif;
}}

/* ── Инфо-блок банка ── */
.bank-info {{
    background: {C_WHITE};
    border: 1px solid {C_BORDER};
    border-left: 5px solid {C_BLUE2};
    border-radius: 0 6px 6px 0;
    padding: 12px 18px;
    margin-bottom: 16px;
    box-shadow: 0 1px 4px rgba(0,43,110,0.1);
    font-family: "Arial", "Helvetica Neue", sans-serif;
}}
.bank-info .name {{
    font-size: 18px;
    font-weight: 700;
    color: #FFFFFF !important;
}}
.bank-info .year {{
    font-size: 13px;
    color: {C_BLUE_LT} !important;
    margin-top: 4px;
}}

/* ── Заголовок таблицы сравнения ── */
.cmp-header {{
    background: {C_BLUE};
    color: #FFFFFF !important;
    font-size: 14px;
    font-weight: 700;
    padding: 10px 16px;
    border-radius: 8px 8px 0 0;
    margin-bottom: 0;
    font-family: "Arial", sans-serif;
}}
</style>
"""

# ─────────────────────────── Константы ──────────────────────────────────

FORM_ORDER = ["balance_sheet", "income_statement", "cash_flow", "equity_changes"]
FORM_NAMES: dict[str, str] = {
    "balance_sheet":    "Отчёт о финансовом положении",
    "income_statement": "Отчёт о прибылях и убытках",
    "cash_flow":        "Отчёт о движении денежных средств",
    "equity_changes":   "Отчёт об изменениях в собственном капитале",
}
FORM_SHORT: dict[str, str] = {
    "balance_sheet":    "Баланс",
    "income_statement": "П&У",
    "cash_flow":        "ДДС",
    "equity_changes":   "Капитал",
}
FORM_ICONS: dict[str, str] = {
    "balance_sheet":    "📊",
    "income_statement": "📈",
    "cash_flow":        "💵",
    "equity_changes":   "🏛",
}


# ─────────────────────────── Утилиты ────────────────────────────────────

import re as _re

# Ключевые показатели МСФО для сравнения
_KPI_PATTERNS: dict[str, list[str]] = {
    "total_assets": [
        r"итого\s+актив",
        r"всего\s+актив",
        r"total\s+assets",
        r"итого\s+по\s+разделу.{0,10}актив",
    ],
    "total_equity": [
        r"итого\s+(?:собственн|капитал)",
        r"всего\s+(?:собственн|капитал)",
        r"total\s+equity",
        r"итого\s+по\s+разделу.{0,15}капитал",
    ],
    "total_liabilities": [
        r"итого\s+обязательств",
        r"всего\s+обязательств",
        r"total\s+liabilities",
    ],
    "net_income": [
        r"чистая\s+прибыл",
        r"прибыл.{1,10}за\s+(?:год|период)",
        r"profit\s+for\s+the\s+year",
        r"net\s+(?:profit|income)",
        r"итого\s+доход",
    ],
    "interest_income": [
        r"процентные\s+доходы",
        r"interest\s+income",
    ],
    "net_interest_income": [
        r"чистые\s+процентные\s+доходы",
        r"net\s+interest\s+income",
    ],
}

_KPI_NAMES: dict[str, str] = {
    "total_assets":      "Итого активы",
    "total_equity":      "Итого капитал",
    "total_liabilities": "Итого обязательства",
    "net_income":        "Чистая прибыль",
    "interest_income":   "Процентные доходы",
    "net_interest_income": "Чистые проц. доходы",
}


def _parse_num(val: str) -> float | None:
    """Конвертирует строку в число (обрабатывает скобки, пробелы, запятые)."""
    v = str(val).strip()
    if not v or v in ("-", "—", "–", "н/д", ""):
        return None
    neg = "(" in v
    v = v.replace("(", "").replace(")", "").replace(" ", "").replace("\u202f", "")
    v = v.replace(",", ".").replace("\xa0", "")
    v = _re.sub(r"[^\d.\-]", "", v)
    try:
        n = float(v)
        return -n if neg else n
    except ValueError:
        return None


def _extract_kpis(df: pd.DataFrame) -> dict[str, float]:
    """Извлекает ключевые МСФО-показатели из DataFrame формы."""
    result: dict[str, float] = {}
    if df.empty or len(df.columns) < 2:
        return result
    for kpi_key, patterns in _KPI_PATTERNS.items():
        for _, row in df.iterrows():
            label = str(row.iloc[0]).strip()
            for pat in patterns:
                if _re.search(pat, label, _re.IGNORECASE):
                    # берём первое числовое значение (обычно текущий год)
                    for cell in row.iloc[1:]:
                        n = _parse_num(str(cell))
                        if n is not None and abs(n) > 0:
                            result[kpi_key] = n
                            break
                    break
    return result


def _fmt_kpi(val: float | None, label: str = "") -> str:
    """Форматирует KPI-значение для отображения."""
    if val is None:
        return "—"
    abs_val = abs(val)
    sign = "-" if val < 0 else ""
    if abs_val >= 1_000_000:
        return f"{sign}{abs_val / 1_000_000:.1f} млн"
    if abs_val >= 1_000:
        return f"{sign}{abs_val / 1_000:.0f} тыс"
    return f"{sign}{abs_val:.1f}"


def _build_combined_excel(bank_data: dict) -> bytes:
    """Единый Excel-файл: все банки на одном листе с разделителями."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl не установлен")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводный отчёт"

    STYLES = {
        "bank": {
            "font":   Font(bold=True, color="FFFFFF", size=13),
            "fill":   PatternFill("solid", fgColor="002B6E"),
            "height": 28,
        },
        "form": {
            "font":   Font(bold=True, color="FFFFFF", size=11),
            "fill":   PatternFill("solid", fgColor="0047AB"),
            "height": 22,
        },
        "col_header": {
            "font":   Font(bold=True, size=10),
            "fill":   PatternFill("solid", fgColor="BDD7EE"),
            "height": 16,
        },
        "section": {
            "font":   Font(bold=True, size=10),
            "fill":   PatternFill("solid", fgColor="D9D9D9"),
            "height": 15,
        },
        "total": {
            "font":   Font(bold=True, size=10),
            "fill":   PatternFill("solid", fgColor="F2F2F2"),
            "height": 14,
        },
        "data": {
            "font":   Font(size=10),
            "fill":   PatternFill("solid", fgColor="FFFFFF"),
            "height": 14,
        },
    }

    _YEAR_RE   = _re.compile(r"\b20\d{2}\b")
    _TOTAL_RE  = _re.compile(r"^(итого|всего|total)\b", _re.IGNORECASE)
    _CAPS_RE   = _re.compile(r"^[А-ЯЁA-Z\s\-/()]{4,}$")

    def _row_style(row_vals: list[str]) -> str:
        label = row_vals[0].strip() if row_vals else ""
        if not label and any(_YEAR_RE.search(c) for c in row_vals[1:] if c.strip()):
            return "col_header"
        if label and _CAPS_RE.match(label) and not any(c.strip() for c in row_vals[1:]):
            return "section"
        if _TOTAL_RE.match(label):
            return "total"
        return "data"

    def _to_num(val: str):
        v = val.strip()
        if not v or v in ("-", "—", "–"):
            return None
        if "%" in v:
            return None
        neg = "(" in v
        v = v.replace("(", "").replace(")", "")
        v = v.replace(" ", "").replace("\u202f", "").replace("\xa0", "")
        v = v.replace(",", ".")
        v = _re.sub(r"[^\d.\-]", "", v)
        try:
            n = float(v)
            n = -n if neg else n
            return int(n) if n == int(n) else round(n, 4)
        except ValueError:
            return None

    MAX_COLS = 1
    cur_row = 1

    for bank_key, data in bank_data.items():
        bank_name = data.get("bank_name", bank_key)
        year      = data.get("year") or "—"

        # ── Шапка банка ───────────────────────────────────────────
        st_b = STYLES["bank"]
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=MAX_COLS)
        cell = ws.cell(row=cur_row, column=1,
                       value=f"  🏦  {bank_name}   |   Год: {year}")
        cell.font      = st_b["font"]
        cell.fill      = st_b["fill"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cur_row].height = st_b["height"]
        cur_row += 1

        for fk in FORM_ORDER:
            r  = data["results"].get(fk, {})
            df = r.get("df", pd.DataFrame())
            if df.empty:
                continue

            n_cols = len(df.columns)
            MAX_COLS = max(MAX_COLS, n_cols)

            # ── Шапка формы ────────────────────────────────────────
            st_f = STYLES["form"]
            ws.merge_cells(start_row=cur_row, start_column=1,
                           end_row=cur_row, end_column=max(n_cols, 2))
            cell = ws.cell(row=cur_row, column=1,
                           value=f"  {FORM_ICONS[fk]}  {FORM_NAMES[fk]}"
                                 f"  ({len(df)} строк)")
            cell.font      = st_f["font"]
            cell.fill      = st_f["fill"]
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[cur_row].height = st_f["height"]
            cur_row += 1

            # ── Строки данных ──────────────────────────────────────
            for _, row_s in df.iterrows():
                vals = [str(v) for v in row_s.tolist()]
                st_key = _row_style(vals)
                st = STYLES[st_key]
                padded = vals + [""] * (MAX_COLS - len(vals))
                for ci, val in enumerate(padded, 1):
                    num  = _to_num(val) if ci > 1 else None
                    cell = ws.cell(row=cur_row, column=ci,
                                   value=num if num is not None else (val or None))
                    cell.font  = st["font"]
                    cell.fill  = st["fill"]
                    cell.alignment = Alignment(
                        horizontal=("right" if num is not None else "left"),
                        vertical="center",
                        indent=(1 if ci == 1 else 0),
                    )
                ws.row_dimensions[cur_row].height = st["height"]
                cur_row += 1

            cur_row += 1  # пустая строка между формами

        cur_row += 2  # пустые строки между банками

    # Ширина колонок
    ws.column_dimensions[get_column_letter(1)].width = 60
    for col in range(2, MAX_COLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@contextlib.contextmanager
def _capture_stdout():
    buf = io.StringIO()
    old = sys.stdout
    sys.stdout = buf
    try:
        yield buf
    finally:
        sys.stdout = old


def _read_df(csv_path: str) -> pd.DataFrame:
    try:
        return pd.read_csv(csv_path, encoding="utf-8-sig", header=None, dtype=str).fillna("")
    except Exception:
        return pd.DataFrame()


def _find_content_list(output_dir: str) -> str | None:
    found = list(Path(output_dir).rglob("*_content_list.json"))
    return str(found[0]) if found else None


def _build_zip(bank_data: dict) -> bytes:
    """Собирает ZIP со всеми результатами всех банков."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for bank_key, data in bank_data.items():
            folder = data["pdf_stem"]
            # XLSX
            if data.get("xlsx_bytes"):
                zf.writestr(f"{folder}/{folder}_ifrs.xlsx", data["xlsx_bytes"])
            # CSV
            for fk in FORM_ORDER:
                df: pd.DataFrame = data["results"].get(fk, {}).get("df", pd.DataFrame())
                if not df.empty:
                    csv_str = df.to_csv(index=False, header=False, encoding="utf-8-sig")
                    zf.writestr(f"{folder}/{folder}_{fk}.csv", csv_str.encode("utf-8-sig"))
    buf.seek(0)
    return buf.read()


# ─────────────────────────── Обработка PDF ──────────────────────────────

def process_pdf(uploaded_file, output_base: str, status_container) -> dict | None:
    """
    Обрабатывает один PDF. Возвращает dict с результатами или None при ошибке.
    status_container — st.empty() для вывода шагов.
    """
    pdf_stem = Path(uploaded_file.name).stem

    # Создаём папку для этого PDF
    out_dir = os.path.join(output_base, pdf_stem)
    os.makedirs(out_dir, exist_ok=True)

    # Сохраняем PDF
    pdf_path = os.path.join(out_dir, uploaded_file.name)
    with open(pdf_path, "wb") as f:
        f.write(uploaded_file.getvalue())

    def _step(text: str):
        status_container.markdown(text)

    _step(f"**⏳ [1/3]** Запуск MinerU — разбор PDF `{uploaded_file.name}`...")

    log_text = ""
    try:
        log_buf = io.StringIO()
        with contextlib.redirect_stdout(log_buf):
            extract_ifrs.run(pdf_path, out_dir)
        log_text = log_buf.getvalue()
    except Exception as exc:
        _step(f"❌ **Ошибка обработки:** {exc}")
        return None

    _step("**⏳ [2/3]** Определение названия банка и года...")

    # Ищем content_list.json для метаданных
    cl_path = _find_content_list(out_dir)
    if cl_path:
        meta = meta_ext.extract_metadata(cl_path, pdf_stem)
    else:
        meta = {"bank_name": pdf_stem, "year": "", "source": "filename"}

    _step("**⏳ [3/3]** Загрузка результатов...")

    # Читаем результаты
    results: dict[str, dict] = {}
    for fk in FORM_ORDER:
        csv_path = os.path.join(out_dir, f"{fk}.csv")
        df = _read_df(csv_path) if os.path.exists(csv_path) else pd.DataFrame()
        results[fk] = {"rows": len(df), "csv": csv_path if df.shape[0] else None, "df": df}

    # Читаем XLSX
    xlsx_path = os.path.join(out_dir, f"{pdf_stem}_ifrs.xlsx")
    xlsx_bytes = open(xlsx_path, "rb").read() if os.path.exists(xlsx_path) else None

    # Читаем единицы измерения (если файл создан extract_ifrs)
    import json as _json
    form_units: dict[str, str] = {}
    units_path = os.path.join(out_dir, "form_units.json")
    if os.path.exists(units_path):
        try:
            with open(units_path, encoding="utf-8") as fu:
                form_units = _json.load(fu)
        except Exception:
            pass

    _step(f"✅ **Готово:** `{uploaded_file.name}`")

    return {
        "pdf_stem":   pdf_stem,
        "filename":   uploaded_file.name,
        "bank_name":  meta["bank_name"],
        "year":       meta["year"],
        "meta_src":   meta["source"],
        "results":    results,
        "xlsx_bytes": xlsx_bytes,
        "log":        log_text,
        "form_units": form_units,
    }


# ─────────────────────────── Компоненты UI ──────────────────────────────

def render_header():
    st.markdown("""
    <div class="app-header">
        <h1>🏦 МСФО Экстрактор</h1>
        <p>Автоматическое извлечение финансовых форм из PDF-отчётов банков</p>
    </div>
    """, unsafe_allow_html=True)


def render_bank_info(data: dict):
    src_badge = {
        "known":    '<span class="badge-ok">✓ Определён точно</span>',
        "legal":    '<span class="badge-warn">~ Из юр. названия</span>',
        "filename": '<span class="badge-err">⚠ Из имени файла</span>',
    }.get(data.get("meta_src", ""), "")

    year_str = f"Год отчётности: <b>{data['year']}</b>" if data.get("year") else \
               '<span style="color:#999">Год не определён</span>'

    # Единицы измерения
    form_units: dict = data.get("form_units", {})
    unique_units = list(dict.fromkeys(v for v in form_units.values() if v))
    if unique_units:
        units_str = " / ".join(unique_units)
        units_html = f'&nbsp;·&nbsp; <span style="color:#0047AB;font-weight:600">Единицы: {units_str}</span>'
    else:
        units_html = ""

    st.markdown(f"""
    <div class="bank-info">
        <div class="name">{data['bank_name']} &nbsp;{src_badge}</div>
        <div class="year">{year_str} &nbsp;·&nbsp; Файл: {data['filename']}{units_html}</div>
    </div>
    """, unsafe_allow_html=True)


def render_stats(results: dict):
    cols = st.columns(4)
    for i, fk in enumerate(FORM_ORDER):
        r = results.get(fk, {})
        rows = r.get("rows", 0)
        found = rows > 0
        badge = '<span class="badge-ok">Найдено</span>' if found \
                else '<span class="badge-err">Не найдено</span>'
        with cols[i]:
            st.markdown(f"""
            <div class="stat-tile">
                <div style="font-size:24px">{FORM_ICONS[fk]}</div>
                <div class="val">{rows if found else "—"}</div>
                <div class="lbl">{FORM_SHORT[fk]}</div>
                <div class="sub">{badge}</div>
            </div>
            """, unsafe_allow_html=True)


def _make_unique_columns(names: list[str]) -> list[str]:
    """
    Делает имена колонок уникальными, добавляя суффикс к дублям.
    Пустые строки заменяет на '_col_N'.
    """
    seen: dict[str, int] = {}
    result: list[str] = []
    for i, name in enumerate(names):
        base = name if name else f"_col_{i}"
        if base in seen:
            seen[base] += 1
            result.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            result.append(base)
    return result


def _smart_columns(df: pd.DataFrame) -> list[str]:
    """
    Определяет умные заголовки: если первая строка содержит годы/«Прим.» —
    использует её как заголовки, иначе генерирует «Показатель / Значение N».
    Гарантирует уникальность всех имён (требование st.dataframe / pyarrow).
    """
    if df.empty:
        return []
    first = [str(v).strip() for v in df.iloc[0]]
    year_re = _re.compile(r"20\d{2}")
    has_year = any(year_re.search(v) for v in first[1:])
    if has_year or any(v.lower() in ("прим.", "прим", "note", "notes") for v in first):
        raw = ["Показатель"] + first[1:]
    else:
        n = len(df.columns)
        raw = ["Показатель"] + [f"Значение {i}" for i in range(1, n)]
    return _make_unique_columns(raw)


def _clean_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Подготавливает DataFrame для отображения:
    - убирает полностью пустые строки
    - присваивает уникальные умные имена колонок
    """
    if df.empty:
        return df
    cols = _smart_columns(df)
    disp = df.copy()
    disp.columns = cols
    # Фильтр: убираем строки, где ВСЕ ячейки пустые
    disp = disp[~disp.apply(lambda r: all(str(v).strip() == "" for v in r), axis=1)]
    return disp.reset_index(drop=True)


def render_forms_tabs(results: dict, key_prefix: str = ""):
    tabs = st.tabs([f"{FORM_ICONS[fk]} {FORM_SHORT[fk]}" for fk in FORM_ORDER])
    for tab, fk in zip(tabs, FORM_ORDER):
        with tab:
            r = results.get(fk, {})
            df: pd.DataFrame = r.get("df", pd.DataFrame())

            st.markdown(f"""
            <div class="form-bar">{FORM_ICONS[fk]} {FORM_NAMES[fk]}
            {"&nbsp;·&nbsp; " + str(len(df)) + " строк" if not df.empty else ""}
            </div>
            """, unsafe_allow_html=True)

            if df.empty:
                st.info("Форма не найдена в документе.")
                continue

            display = _clean_display_df(df)

            # Ключевые показатели над таблицей
            kpis = _extract_kpis(df)
            if kpis:
                kpi_cols = st.columns(min(len(kpis), 4))
                shown_keys = list(kpis.keys())[:4]
                for ki, kkey in enumerate(shown_keys):
                    with kpi_cols[ki]:
                        st.metric(
                            label=_KPI_NAMES.get(kkey, kkey),
                            value=_fmt_kpi(kpis[kkey]),
                        )

            st.dataframe(
                display,
                use_container_width=True,
                hide_index=True,
                height=min(620, 36 * len(display) + 40),
                key=f"tbl_{key_prefix}_{fk}",
            )


def render_download(data: dict, key_sfx: str = ""):
    st.markdown("#### ⬇ Скачать")
    col1, col2 = st.columns([1, 1])

    with col1:
        if data.get("xlsx_bytes"):
            st.download_button(
                label="📊 Excel (все формы)",
                data=data["xlsx_bytes"],
                file_name=f"{data['pdf_stem']}_ifrs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_xl_{key_sfx}",
            )
        else:
            st.warning("Excel не создан")

    with col2:
        for fk in FORM_ORDER:
            r = data["results"].get(fk, {})
            df: pd.DataFrame = r.get("df", pd.DataFrame())
            if not df.empty:
                csv_bytes = df.to_csv(index=False, header=False).encode("utf-8-sig")
                st.download_button(
                    label=f"📄 CSV · {FORM_SHORT[fk]}",
                    data=csv_bytes,
                    file_name=f"{data['pdf_stem']}_{fk}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key=f"dl_csv_{key_sfx}_{fk}",
                )


def render_comparison(bank_data: dict):
    """Сводная таблица сравнения нескольких банков — по строкам и KPI."""
    st.markdown('<div class="cmp-header">📊 Сравнение отчётов по банкам</div>',
                unsafe_allow_html=True)

    # ── Таблица 1: количество найденных строк ─────────────────────────
    st.markdown("**Количество найденных строк по формам**")
    rows_count = []
    for bank_key, data in bank_data.items():
        row: dict = {
            "Банк": data["bank_name"],
            "Год":  data["year"] or "—",
        }
        total_rows = 0
        for fk in FORM_ORDER:
            n = data["results"].get(fk, {}).get("rows", 0)
            row[FORM_SHORT[fk]] = n if n > 0 else 0
            total_rows += n
        row["Итого строк"] = total_rows
        rows_count.append(row)

    df_count = pd.DataFrame(rows_count)
    st.dataframe(
        df_count,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Итого строк": st.column_config.NumberColumn(format="%d", help="Сумма строк по всем формам"),
            **{FORM_SHORT[fk]: st.column_config.NumberColumn(format="%d") for fk in FORM_ORDER},
        },
    )

    # ── Таблица 2: ключевые МСФО-показатели ──────────────────────────
    st.markdown("**Ключевые показатели МСФО**")
    kpi_rows = []
    for bank_key, data in bank_data.items():
        krow: dict = {
            "Банк": data["bank_name"],
            "Год":  data["year"] or "—",
        }
        # Собираем KPI из всех форм
        all_kpis: dict[str, float] = {}
        for fk in FORM_ORDER:
            df_form = data["results"].get(fk, {}).get("df", pd.DataFrame())
            all_kpis.update(_extract_kpis(df_form))

        for kk, kname in _KPI_NAMES.items():
            val = all_kpis.get(kk)
            krow[kname] = _fmt_kpi(val) if val is not None else "—"
        kpi_rows.append(krow)

    df_kpi = pd.DataFrame(kpi_rows)
    st.dataframe(df_kpi, use_container_width=True, hide_index=True)


def render_export_section(bank_data: dict):
    """Секция экспорта."""
    st.markdown("---")
    st.markdown("### 📥 Экспорт результатов")

    with st.expander("⚙ Настройки экспорта", expanded=True):
        col_left, col_right = st.columns(2)

        with col_left:
            report_type = st.radio(
                "Тип отчёта:",
                [
                    "📋 Индивидуальный по банку",
                    "📊 Сводная таблица (все банки, 1 лист)",
                    "📦 Полный архив (ZIP)",
                ],
                index=0,
            )

        with col_right:
            bank_keys = list(bank_data.keys())
            if report_type == "📋 Индивидуальный по банку" and len(bank_keys) > 1:
                selected_keys = st.multiselect(
                    "Банки:",
                    options=bank_keys,
                    format_func=lambda k: bank_data[k]["bank_name"],
                    default=bank_keys,
                )
            else:
                selected_keys = bank_keys

            is_zip     = report_type == "📦 Полный архив (ZIP)"
            is_combined = report_type == "📊 Сводная таблица (все банки, 1 лист)"
            format_opts = st.multiselect(
                "Форматы:",
                [".xlsx", ".csv"],
                default=[".xlsx"],
                disabled=(is_zip or is_combined),
            )

    if st.button("📊 Сгенерировать", type="primary"):
        if is_zip:
            zip_bytes = _build_zip({k: bank_data[k] for k in selected_keys})
            st.download_button(
                "⬇ Скачать ZIP-архив",
                data=zip_bytes,
                file_name="ifrs_export.zip",
                mime="application/zip",
                key="dl_zip_final",
            )

        elif is_combined:
            with st.spinner("Формирование сводного Excel..."):
                try:
                    combined_bytes = _build_combined_excel(
                        {k: bank_data[k] for k in selected_keys}
                    )
                    st.download_button(
                        "⬇ Скачать сводный Excel (все банки, 1 лист)",
                        data=combined_bytes,
                        file_name="all_banks_combined.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_combined_xl",
                        use_container_width=True,
                    )
                    st.success("✅ Сводный Excel сформирован!")
                except Exception as e:
                    st.error(f"Ошибка формирования Excel: {e}")

        else:  # индивидуальный
            for k in selected_keys:
                d = bank_data[k]
                st.markdown(f"**{d['bank_name']}** ({d['year'] or '—'})")
                _render_individual_exports(d, k, format_opts)


def _render_individual_exports(data: dict, key_sfx: str, format_opts: list):
    cols = st.columns(len(FORM_ORDER) + 1)
    with cols[0]:
        if ".xlsx" in format_opts and data.get("xlsx_bytes"):
            st.download_button(
                "📊 XLSX",
                data["xlsx_bytes"],
                f"{data['pdf_stem']}_ifrs.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"exp_xl_{key_sfx}",
                use_container_width=True,
            )
    if ".csv" in format_opts:
        for i, fk in enumerate(FORM_ORDER):
            df: pd.DataFrame = data["results"].get(fk, {}).get("df", pd.DataFrame())
            with cols[i + 1]:
                if not df.empty:
                    st.download_button(
                        f"{FORM_SHORT[fk]}",
                        df.to_csv(index=False, header=False).encode("utf-8-sig"),
                        f"{data['pdf_stem']}_{fk}.csv",
                        "text/csv",
                        key=f"exp_csv_{key_sfx}_{fk}",
                        use_container_width=True,
                    )
                else:
                    st.button(f"✗ {FORM_SHORT[fk]}", disabled=True,
                              key=f"exp_na_{key_sfx}_{fk}", use_container_width=True)


# ─────────────────────────── Сайдбар ────────────────────────────────────

def render_sidebar() -> str:
    with st.sidebar:
        st.markdown(f"""
        <div style="background:{C_BLUE};color:white;padding:14px 16px;
                    border-radius:8px;margin-bottom:16px">
            <b style="font-size:15px">⚙ Параметры</b>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("**Метод разбора PDF**")
        parse_method = st.selectbox(
            "Метод:",
            ["txt — текстовый PDF (быстро)", "ocr — сканированный PDF (медленно)"],
            label_visibility="collapsed",
        )

        st.markdown("---")
        st.markdown("**Формы МСФО**")
        for fk in FORM_ORDER:
            st.markdown(f"<small>{FORM_ICONS[fk]} {FORM_NAMES[fk]}</small>",
                        unsafe_allow_html=True)

        st.markdown("---")
        if st.button("🗑 Очистить все данные", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

        st.markdown("---")
        st.caption("Использует MinerU · openpyxl")

    return "ocr" if "ocr" in parse_method else "txt"


# ─────────────────────────────── Main ───────────────────────────────────

def main():
    st.set_page_config(
        page_title="МСФО Экстрактор",
        page_icon="🏦",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    st.markdown(CSS, unsafe_allow_html=True)

    render_header()
    render_sidebar()

    # ── Постоянная папка для сессии ──────────────────────────────────
    if "session_dir" not in st.session_state:
        st.session_state.session_dir = tempfile.mkdtemp(prefix="ifrs_")
    if "bank_data" not in st.session_state:
        st.session_state.bank_data = {}  # key=pdf_stem → данные

    session_dir = st.session_state.session_dir
    bank_data: dict = st.session_state.bank_data

    # ── Загрузчик файлов ─────────────────────────────────────────────
    st.markdown("### 📂 Загрузка файлов")
    uploaded_files = st.file_uploader(
        "Перетащите PDF-отчёты или нажмите «Browse files»",
        type=["pdf"],
        accept_multiple_files=True,
        help="Поддерживается загрузка нескольких банковских отчётов",
        label_visibility="collapsed",
    )

    if not uploaded_files and not bank_data:
        st.markdown("""
        <div class="card" style="text-align:center;padding:36px;color:#5A6478">
            <div style="font-size:48px">📄</div>
            <div style="font-size:16px;font-weight:600;margin-top:8px">
                Загрузите PDF-отчёты для анализа
            </div>
            <div style="font-size:13px;margin-top:6px">
                Поддерживаются МСФО-отчёты российских банков в формате PDF
            </div>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Кнопка запуска ──────────────────────────────────────────────
    col_run, col_space = st.columns([2, 3])
    with col_run:
        run_clicked = st.button("🚀 Извлечь данные", type="primary",
                                use_container_width=True,
                                disabled=not uploaded_files)

    # ── Обработка ───────────────────────────────────────────────────
    if run_clicked and uploaded_files:
        progress = st.progress(0)
        total = len(uploaded_files)

        for idx, uf in enumerate(uploaded_files):
            stem = Path(uf.name).stem

            with st.container():
                st.markdown(f"""
                <div class="card">
                <div class="card-header">Обработка файла {idx+1}/{total}: {uf.name}</div>
                """, unsafe_allow_html=True)
                status_el = st.empty()

                result = process_pdf(uf, session_dir, status_el)

                if result:
                    bank_data[stem] = result
                    status_el.markdown(
                        f'✅ Завершено — найдено форм: '
                        f'**{sum(1 for r in result["results"].values() if r["rows"] > 0)}/4**'
                    )
                else:
                    status_el.error("Обработка не выполнена")

                st.markdown("</div>", unsafe_allow_html=True)

            progress.progress((idx + 1) / total)

        progress.progress(1.0)
        st.success(f"✅ Обработано файлов: {total}")

    # ── Результаты ──────────────────────────────────────────────────
    if not bank_data:
        return

    st.markdown("---")
    st.markdown("### 📋 Результаты")

    # Сравнение если несколько банков
    if len(bank_data) > 1:
        render_comparison(bank_data)
        st.markdown("---")

    # Детали по каждому банку
    if len(bank_data) > 1:
        bank_tabs = st.tabs([
            f"{FORM_ICONS['balance_sheet']} {d['bank_name']} {d['year'] or ''}"
            for d in bank_data.values()
        ])
    else:
        bank_tabs = [st.container()]

    for tab, (stem, data) in zip(bank_tabs, bank_data.items()):
        with tab:
            render_bank_info(data)
            render_stats(data["results"])
            render_forms_tabs(data["results"], key_prefix=stem)
            st.markdown("---")
            render_download(data, key_sfx=stem)

            with st.expander("📋 Журнал обработки"):
                log = data.get("log", "")
                st.code(log if log else "(нет данных)", language=None)

    # Экспорт
    render_export_section(bank_data)

    # Сброс сессии
    st.markdown("---")
    col_fin, _ = st.columns([1, 3])
    with col_fin:
        if st.button("🔄 Начать новый анализ", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


if __name__ == "__main__":
    main()
